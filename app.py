import csv
import io
import os
import re
import sys
import time
from pathlib import Path

import serial
import serial.tools.list_ports
import webview
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

import jt_parser


BAUD_RATE = 115200


def resource_path(relative):
    """Resolve a path to a bundled resource.

    When frozen by PyInstaller, data files live in the temporary folder
    referenced by sys._MEIPASS. In normal development we fall back to the
    directory of this script.
    """
    base = getattr(
        sys,
        "_MEIPASS",
        os.path.dirname(os.path.abspath(__file__)),
    )
    return os.path.join(base, relative)


class API:
    def __init__(self):
        self.port = None

    # =========================================================
    # Serial / device
    # =========================================================

    def find_esp32(self):
        for port in serial.tools.list_ports.comports():
            description = (port.description or "").lower()

            if (
                "cp210" in description
                or "silicon labs" in description
                or "ch340" in description
                or "usb serial" in description
            ):
                self.port = port.device

                return {
                    "success": True,
                    "port": port.device,
                    "description": port.description or "Unknown",
                    "manufacturer": port.manufacturer or "Unknown",
                    "serial_number": port.serial_number or "Unknown",
                }

        self.port = None

        return {
            "success": False,
            "message": "ESP32 not found.",
        }

    def open_serial(self):
        if self.port is None:
            result = self.find_esp32()

            if not result["success"]:
                raise RuntimeError("ESP32 not found.")

        ser = serial.Serial(
            self.port,
            BAUD_RATE,
            timeout=0.5,
            write_timeout=2,
        )

        # CP210x boards commonly reset when the serial port opens.
        time.sleep(1.5)
        ser.reset_input_buffer()

        return ser

    def wait_response(self, ser, timeout=5):
        start = time.time()

        while time.time() - start <= timeout:
            raw = ser.readline()

            if not raw:
                continue

            return raw.decode("utf-8", errors="ignore").strip()

        return "TIMEOUT"

    # =========================================================
    # JT parsing
    # =========================================================

    def choose_jt_pdf(self):
        result = window.create_file_dialog(
            webview.FileDialog.OPEN,
            file_types=("PDF Files (*.pdf)",),
        )

        if not result:
            return {
                "success": False,
                "cancelled": True,
            }

        path = result[0] if isinstance(result, (tuple, list)) else result

        try:
            jt_name, steps = jt_parser.parse_document(path)

            if not steps:
                return {
                    "success": False,
                    "message": "No JT steps were found in this PDF.",
                }

            return {
                "success": True,
                "filename": os.path.basename(path),
                "jt_name": jt_name,
                "steps": [
                    {
                        "number": number,
                        "description": description,
                    }
                    for number, description in steps.items()
                ],
            }

        except Exception as exc:
            return {
                "success": False,
                "message": str(exc),
            }

    # =========================================================
    # Create study
    # =========================================================

    @staticmethod
    def _sanitize_field(value):
        """Keep a filename component filesystem- and serial-protocol-safe."""
        value = (value or "").strip()

        # Collapse whitespace to nothing, then drop anything that is not a
        # letter, digit, dash, or dot. Part numbers like "172-0002" survive.
        value = re.sub(r"\s+", "", value)
        value = re.sub(r"[^A-Za-z0-9.-]", "", value)

        return value

    @staticmethod
    def _sanitize_name(value):
        """Sanitize an optional custom study name into a safe base filename."""
        value = (value or "").strip()

        # Spaces become underscores so the name reads well but stays safe for
        # the serial protocol; drop anything else that isn't allowed.
        value = re.sub(r"\s+", "_", value)
        value = re.sub(r"[^A-Za-z0-9._-]", "", value)

        # Strip a trailing .csv (any case) so we can re-add it consistently.
        if value.lower().endswith(".csv"):
            value = value[:-4]

        return value

    def create_study(self, part_number, work_order, custom_name, steps):
        part_number = self._sanitize_field(part_number)
        work_order = self._sanitize_field(work_order)
        custom_name = self._sanitize_name(custom_name)

        if not part_number:
            return {
                "success": False,
                "message": "Part number is required.",
            }

        if not work_order:
            return {
                "success": False,
                "message": "Work order number is required.",
            }

        if not steps:
            return {
                "success": False,
                "message": "No parsed steps are available.",
            }

        # A custom name, when supplied, overrides the default convention.
        # Default naming convention: <part number>_<work order number>.csv
        if custom_name:
            filename = f"{custom_name}.csv"
        else:
            filename = f"{part_number}_{work_order}.csv"

        ser = None

        try:
            ser = self.open_serial()

            ser.write(f"NEWSTUDY {filename}\n".encode("utf-8"))
            ser.flush()

            response = self.wait_response(ser)

            if response == "ERROR:FILE_ALREADY_EXISTS":
                return {
                    "success": False,
                    "message": f"{filename} already exists on the ESP32.",
                }

            if response != "OK:STUDY_CREATED":
                return {
                    "success": False,
                    "message": f"ESP32 returned: {response}",
                }

            for step in steps:
                number = step["number"]

                description = str(step.get("description", ""))
                description = (
                    description.replace("|", "/")
                    .replace("\r", " ")
                    .replace("\n", " ")
                )

                command = f"STEP {filename}|{number}|{description}\n"

                ser.write(command.encode("utf-8"))
                ser.flush()

                response = self.wait_response(ser)

                if response != "OK:STEP_ADDED":
                    return {
                        "success": False,
                        "message": (
                            f"The study file was created, but step {number} "
                            f"failed to upload. ESP32 returned: {response}"
                        ),
                    }

            return {
                "success": True,
                "filename": filename,
                "step_count": len(steps),
            }

        except Exception as exc:
            return {
                "success": False,
                "message": str(exc),
            }

        finally:
            if ser is not None and ser.is_open:
                ser.close()

    # =========================================================
    # List CSVs
    # =========================================================

    def get_csv_files(self):
        ser = None

        try:
            ser = self.open_serial()
            ser.write(b"LIST\n")
            ser.flush()

            files = []
            receiving = False
            start = time.time()

            while time.time() - start <= 5:
                raw = ser.readline()

                if not raw:
                    continue

                line = raw.decode("utf-8", errors="ignore").strip()

                if line == "LIST_START":
                    receiving = True
                    continue

                if line == "LIST_END":
                    return {
                        "success": True,
                        "files": files,
                        "count": len(files),
                    }

                if receiving and line.endswith(".csv"):
                    files.append(line.lstrip("/"))

            return {
                "success": False,
                "message": "Timed out while reading the ESP32 file list.",
            }

        except Exception as exc:
            return {
                "success": False,
                "message": str(exc),
            }

        finally:
            if ser is not None and ser.is_open:
                ser.close()

    # =========================================================
    # Read CSV
    # =========================================================

    def read_csv(self, filename):
        ser = None

        try:
            ser = self.open_serial()
            ser.write(f"GET {filename}\n".encode("utf-8"))
            ser.flush()

            lines = []
            receiving = False
            start = time.time()

            while time.time() - start <= 8:
                raw = ser.readline()

                if not raw:
                    continue

                line = raw.decode(
                    "utf-8",
                    errors="ignore",
                ).rstrip("\r\n")

                if line == "ERROR:FILE_NOT_FOUND":
                    return {
                        "success": False,
                        "message": "File not found on the ESP32.",
                    }

                if line == "FILE_START":
                    receiving = True
                    continue

                if line == "FILE_END":
                    reader = csv.reader(io.StringIO("\n".join(lines)))
                    rows = list(reader)

                    if not rows:
                        return {
                            "success": True,
                            "filename": filename,
                            "headers": [],
                            "rows": [],
                            "row_count": 0,
                        }

                    return {
                        "success": True,
                        "filename": filename,
                        "headers": rows[0],
                        "rows": rows[1:],
                        "row_count": len(rows) - 1,
                    }

                if receiving:
                    lines.append(line)

            return {
                "success": False,
                "message": "Timed out while reading the CSV.",
            }

        except Exception as exc:
            return {
                "success": False,
                "message": str(exc),
            }

        finally:
            if ser is not None and ser.is_open:
                ser.close()

    # =========================================================
    # Download raw CSV
    # =========================================================

    def download_csv(self, filename):
        result = self.read_csv(filename)

        if not result["success"]:
            return result

        path = window.create_file_dialog(
            webview.FileDialog.SAVE,
            save_filename=filename,
            file_types=("CSV Files (*.csv)",),
        )

        if not path:
            return {
                "success": False,
                "cancelled": True,
            }

        destination = path[0] if isinstance(path, (tuple, list)) else path

        try:
            if not str(destination).lower().endswith(".csv"):
                destination = str(destination) + ".csv"

            with open(
                destination,
                "w",
                newline="",
                encoding="utf-8",
            ) as file:
                writer = csv.writer(file)
                writer.writerow(result["headers"])
                writer.writerows(result["rows"])

            return {
                "success": True,
                "path": str(destination),
            }

        except Exception as exc:
            return {
                "success": False,
                "message": str(exc),
            }

    # =========================================================
    # Delete CSV
    # =========================================================

    def delete_csv(self, filename):
        ser = None

        try:
            ser = self.open_serial()
            ser.write(f"DELETE {filename}\n".encode("utf-8"))
            ser.flush()

            response = self.wait_response(ser)

            if response.startswith("OK:DELETED:"):
                return {"success": True}

            if response == "ERROR:FILE_NOT_FOUND":
                return {
                    "success": False,
                    "message": "The file no longer exists on the ESP32.",
                }

            if response == "ERROR:DELETE_FAILED":
                return {
                    "success": False,
                    "message": "The ESP32 could not delete the file.",
                }

            return {
                "success": False,
                "message": f"ESP32 returned: {response}",
            }

        except Exception as exc:
            return {
                "success": False,
                "message": str(exc),
            }

        finally:
            if ser is not None and ser.is_open:
                ser.close()

    # =========================================================
    # Excel export
    # =========================================================

    def export_excel(self, filenames):
        if not filenames:
            return {
                "success": False,
                "message": "Select at least one CSV file.",
            }

        path = window.create_file_dialog(
            webview.FileDialog.SAVE,
            save_filename="ESP32_Studies.xlsx",
            file_types=("Excel Workbook (*.xlsx)",),
        )

        if not path:
            return {
                "success": False,
                "cancelled": True,
            }

        destination = path[0] if isinstance(path, (tuple, list)) else path

        if not str(destination).lower().endswith(".xlsx"):
            destination = str(destination) + ".xlsx"

        try:
            workbook = Workbook()
            workbook.remove(workbook.active)

            for filename in filenames:
                result = self.read_csv(filename)

                if not result["success"]:
                    continue

                base_name = Path(filename).stem[:31] or "Study"
                sheet_name = base_name
                counter = 2

                while sheet_name in workbook.sheetnames:
                    suffix = f"_{counter}"
                    sheet_name = base_name[: 31 - len(suffix)] + suffix
                    counter += 1

                sheet = workbook.create_sheet(sheet_name)

                for column, header in enumerate(result["headers"], start=1):
                    cell = sheet.cell(
                        row=1,
                        column=column,
                        value=header,
                    )
                    cell.font = Font(bold=True)

                for row_number, row in enumerate(result["rows"], start=2):
                    for column, value in enumerate(row, start=1):
                        sheet.cell(
                            row=row_number,
                            column=column,
                            value=value,
                        )

                sheet.freeze_panes = "A2"

                if result["headers"]:
                    sheet.auto_filter.ref = sheet.dimensions

                # Fixed export layout so long step descriptions stay
                # contained inside the Step Description column.
                sheet.column_dimensions["A"].width = 14
                sheet.column_dimensions["B"].width = 65
                sheet.column_dimensions["C"].width = 16

                for cell in sheet[1]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(vertical="center")

                for row in range(2, sheet.max_row + 1):
                    sheet.cell(row=row, column=1).alignment = Alignment(
                        vertical="top"
                    )
                    sheet.cell(row=row, column=2).alignment = Alignment(
                        wrap_text=True,
                        vertical="top"
                    )
                    sheet.cell(row=row, column=3).alignment = Alignment(
                        horizontal="center",
                        vertical="top"
                    )

                    # Keep exported rows compact while preventing text
                    # from visually overflowing into adjacent cells.
                    sheet.row_dimensions[row].height = 30

            if not workbook.sheetnames:
                return {
                    "success": False,
                    "message": "No valid CSV files could be exported.",
                }

            workbook.save(destination)

            return {
                "success": True,
                "path": str(destination),
            }

        except Exception as exc:
            return {
                "success": False,
                "message": str(exc),
            }


api = API()

window = webview.create_window(
    "ESP32 Study Manager",
    resource_path("web/index.html"),
    js_api=api,
    width=1150,
    height=780,
    min_size=(900, 620),
)

webview.start()
